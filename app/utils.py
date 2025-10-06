from django.template.loader import render_to_string
from weasyprint import HTML
from django.conf import settings
from django.core.mail import EmailMessage
import tempfile
from django.db.models import Sum
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from app.administration.models import TeacherPayment, Invoice, Expense, FinancialReport
from app.administration.serializers import ExpenseSerializer, FinancialReportSerializer

def render_to_pdf(template_src, context_dict):
    html_string = render_to_string(template_src, context_dict)
    html = HTML(string=html_string, base_url=settings.BASE_DIR)

    # Создаем временный файл без блокировки
    result = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    html.write_pdf(target=result.name)
    result.seek(0)
    return result




def _to_bytes(pdf_file):
    """Преобразует результат render_to_pdf в bytes"""
    if hasattr(pdf_file, "read"):
        pdf_file.seek(0)
        return pdf_file.read()
    return pdf_file


def send_financial_reports_to_manager():
    attachments = []

    # ===== 1. Расчёты с преподавателями =====
    payments = TeacherPayment.objects.all()
    total_amount = payments.aggregate(Sum("paid_amount"))["paid_amount__sum"] or 0
    context = {"payments": payments, "total_amount": total_amount}
    pdf_file = render_to_pdf("reports/teacher_payments.html", context)
    attachments.append(("teacher_payments.pdf", _to_bytes(pdf_file), "application/pdf"))

    # ===== 2. Доходы =====
    incomes = Invoice.objects.all().select_related("direction", "student", "group")
    total_amount = incomes.aggregate(Sum("amount"))["amount__sum"] or 0
    context = {"incomes": incomes, "total_amount": total_amount}
    pdf_file = render_to_pdf("reports/income_pdf_template.html", context)
    attachments.append(("incomes.pdf", _to_bytes(pdf_file), "application/pdf"))

    # ===== 3. Расходы =====
    expenses = Expense.objects.all().select_related("teacher")
    serializer = ExpenseSerializer(expenses, many=True)
    total_amount = expenses.aggregate(Sum("amount"))["amount__sum"] or 0
    context = {"expenses": serializer.data, "total_amount": total_amount}
    pdf_file = render_to_pdf("reports/expense_pdf_template.html", context)
    attachments.append(("expenses.pdf", _to_bytes(pdf_file), "application/pdf"))

    # ===== 4. Финансовый результат =====
    reports = FinancialReport.objects.all()
    serializer = FinancialReportSerializer(reports, many=True)
    context = {"reports": serializer.data}
    pdf_file = render_to_pdf("reports/financial_report_pdf_template.html", context)
    attachments.append(("financial_report.pdf", _to_bytes(pdf_file), "application/pdf"))

    # ===== Отправка письма =====
    subject = "Финансовые отчёты"
    body = (
        "Добрый день!\n\n"
        "Во вложении актуальные финансовые отчёты:\n"
        "— Расчёты с преподавателями\n"
        "— Доходы\n"
        "— Расходы\n"
        "— Финансовый результат\n\n"
        "С уважением,\nАвтоматизированная система"
    )

    email = EmailMessage(
        subject=subject,
        body=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[settings.MANAGER_EMAIL],
    )

    for filename, content, mimetype in attachments:
        email.attach(filename, content, mimetype)

    email.send()
    return True


def generate_excel(filename: str, headers: list, rows: list, title: str = None, total: float = None):
    """
    Универсальная генерация Excel-файла.
    :param filename: имя файла для сохранения (например, "income_report.xlsx")
    :param headers: список заголовков таблицы
    :param rows: список строк (list of lists), каждая строка — это список значений
    :param title: (опционально) заголовок на первой строке
    :param total: (опционально) итоговая сумма
    :return: HttpResponse с xlsx файлом
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Заголовок
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(size=14, bold=True)
        ws.append([])

    # Шапка таблицы
    ws.append(headers)

    # Данные
    for row in rows:
        ws.append(row)

    # Итог
    if total is not None:
        ws.append([])
        ws.append(["Итого", total])

    # Возврат HttpResponse
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



def generate_invoice_pdf(payment):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Заголовок
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "American Dream")

    # Чек и дата
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 80, f"Чек #{payment.id}")
    p.drawString(250, height - 80, payment.date.strftime("%d.%m.%Y %H:%M"))

    # Детали
    p.drawString(50, height - 120, "Наименование")
    p.drawString(300, height - 120, "Сумма")
    
    p.drawString(50, height - 140, f"{payment.invoice.months.title} {payment.invoice.student.get_full_name()}")
    p.drawString(300, height - 140, f"{payment.total_amount} KGS")

    # Способ оплаты
    method = "Наличные" if payment.cash_amount else "Перевод" if payment.transfer_amount else "Онлайн"
    p.drawString(50, height - 180, f"Способ оплаты: {method}")
    p.drawString(50, height - 200, f"Клиент: {payment.invoice.student.get_full_name()}")

    # Итог
    p.drawString(50, height - 230, f"Итого: {payment.total_amount} KGS")

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer